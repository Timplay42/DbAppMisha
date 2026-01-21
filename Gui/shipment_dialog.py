# Gui/shipment_dialog.py
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel,
    QLineEdit, QPushButton, QDoubleSpinBox, QDateTimeEdit,
    QMessageBox, QComboBox, QGroupBox, QGridLayout,
    QTextEdit, QRadioButton, QButtonGroup, QFileDialog
)
from PySide6.QtCore import QDateTime, Qt
from PySide6.QtGui import QFont
import datetime
import os


class ShipmentDialog(QDialog):
    """Диалог для создания/редактирования перевозки"""

    def __init__(self, parent=None, shipment=None,
                 available_cars=None, available_drivers=None,
                 available_routes=None, available_tariffs=None):
        super().__init__(parent)
        self.setWindowTitle("Редактирование перевозки" if shipment else "Создание перевозки")
        self.setMinimumWidth(800)

        self.shipment = shipment
        self.available_cars = available_cars or []
        self.available_drivers = available_drivers or []
        self.available_routes = available_routes or []
        self.available_tariffs = available_tariffs or []

        # Виджеты
        self.shipment_date = QDateTimeEdit()
        self.shipment_date.setCalendarPopup(True)
        self.shipment_date.setDateTime(QDateTime.currentDateTime())
        self.shipment_date.setDisplayFormat("dd.MM.yyyy HH:mm")

        self.cargo_weight = QDoubleSpinBox()
        self.cargo_weight.setRange(0.1, 100000)
        self.cargo_weight.setSuffix(" кг")
        self.cargo_weight.setDecimals(1)

        self.status = QComboBox()
        self.status.addItems(["pending", "in_transit", "delivered", "cancelled"])
        # Устанавливаем читаемые названия статусов
        self.status.setItemText(0, "⏳ Ожидает")
        self.status.setItemText(1, "🚛 В пути")
        self.status.setItemText(2, "✅ Доставлено")
        self.status.setItemText(3, "❌ Отменено")

        self.car_combo = QComboBox()
        self.car_combo.addItem("Выберите автомобиль", None)
        for car in self.available_cars:
            self.car_combo.addItem(car["full_info"], car["id"])

        self.driver_combo = QComboBox()
        self.driver_combo.addItem("Выберите водителя", None)
        for driver in self.available_drivers:
            text = f"{driver['full_name']}"
            if driver.get('license_number'):
                text += f" ({driver['license_number']})"
            self.driver_combo.addItem(text, driver["id"])

        self.route_combo = QComboBox()
        self.route_combo.addItem("Выберите маршрут", None)
        for route in self.available_routes:
            info = f"{route['origin']} → {route['destination']} ({route['distance_km']} км)"
            self.route_combo.addItem(info, route["id"])

        self.tariff_combo = QComboBox()
        self.tariff_combo.addItem("Выберите тариф", None)
        for tariff in self.available_tariffs:
            self.tariff_combo.addItem(tariff["full_info"], tariff["id"])

        # Расчетные поля
        self.distance_label = QLabel("0 км")
        self.price_per_km_label = QLabel("0 руб/км")
        self.min_price_label = QLabel("0 руб")
        self.total_cost_label = QLabel("0 руб")
        self.total_cost_label.setFont(QFont("Arial", 12, QFont.Bold))
        self.total_cost_label.setStyleSheet("color: green;")

        # Кнопки
        self.calculate_btn = QPushButton("📊 Рассчитать стоимость")
        self.calculate_btn.clicked.connect(self.calculate_cost)

        self.export_btn = QPushButton("📈 Экспорт перевозок")
        self.export_btn.clicked.connect(self.show_export_dialog)

        self.save_btn = QPushButton("💾 Сохранить")
        self.save_btn.clicked.connect(self.validate_and_accept)

        self.cancel_btn = QPushButton("❌ Отмена")
        self.cancel_btn.clicked.connect(self.reject)

        self.setup_ui()

        if shipment:
            self.load_data(shipment)

        # Подключаем сигналы
        self.cargo_weight.valueChanged.connect(self.check_car_capacity)
        self.car_combo.currentIndexChanged.connect(self.check_car_capacity)
        self.route_combo.currentIndexChanged.connect(self.calculate_cost)
        self.tariff_combo.currentIndexChanged.connect(self.calculate_cost)

    def setup_ui(self):
        """Настройка интерфейса"""
        layout = QVBoxLayout(self)

        # Основная информация
        info_group = QGroupBox("Основная информация")
        info_layout = QGridLayout()

        info_layout.addWidget(QLabel("Дата и время *:"), 0, 0)
        info_layout.addWidget(self.shipment_date, 0, 1)

        info_layout.addWidget(QLabel("Вес груза *:"), 1, 0)
        info_layout.addWidget(self.cargo_weight, 1, 1)

        info_layout.addWidget(QLabel("Статус *:"), 3, 0)
        info_layout.addWidget(self.status, 3, 1)

        info_group.setLayout(info_layout)
        layout.addWidget(info_group)

        # Назначения
        assign_group = QGroupBox("Назначения")
        assign_layout = QGridLayout()

        assign_layout.addWidget(QLabel("Автомобиль *:"), 0, 0)
        assign_layout.addWidget(self.car_combo, 0, 1)

        assign_layout.addWidget(QLabel("Водитель *:"), 1, 0)
        assign_layout.addWidget(self.driver_combo, 1, 1)

        assign_layout.addWidget(QLabel("Маршрут *:"), 2, 0)
        assign_layout.addWidget(self.route_combo, 2, 1)

        assign_layout.addWidget(QLabel("Тариф *:"), 3, 0)
        assign_layout.addWidget(self.tariff_combo, 3, 1)

        assign_group.setLayout(assign_layout)
        layout.addWidget(assign_group)

        # Расчет стоимости
        calc_group = QGroupBox("Расчет стоимости")
        calc_layout = QGridLayout()

        calc_layout.addWidget(QLabel("Расстояние:"), 0, 0)
        calc_layout.addWidget(self.distance_label, 0, 1)

        calc_layout.addWidget(QLabel("Цена за км:"), 1, 0)
        calc_layout.addWidget(self.price_per_km_label, 1, 1)

        calc_layout.addWidget(QLabel("Минимальная цена:"), 2, 0)
        calc_layout.addWidget(self.min_price_label, 2, 1)

        calc_layout.addWidget(QLabel("Итоговая стоимость:"), 3, 0)
        calc_layout.addWidget(self.total_cost_label, 3, 1)

        calc_layout.addWidget(self.calculate_btn, 4, 0, 1, 2)

        calc_group.setLayout(calc_layout)
        layout.addWidget(calc_group)

        # Кнопки
        btn_layout = QHBoxLayout()
        btn_layout.addWidget(self.export_btn)
        btn_layout.addStretch()
        btn_layout.addWidget(self.save_btn)
        btn_layout.addWidget(self.cancel_btn)

        layout.addLayout(btn_layout)

    def load_data(self, shipment):
        """Загрузить данные перевозки"""
        # Дата
        if shipment.get("shipment_date"):
            dt = QDateTime.fromString(shipment["shipment_date"], Qt.ISODate)
            self.shipment_date.setDateTime(dt)

        # Вес и тип груза
        self.cargo_weight.setValue(shipment.get("cargo_weight", 0))

        # Статус
        status = shipment.get("status", "pending")
        # Преобразуем статус в читаемый вид
        status_mapping = {
            "pending": "⏳ Ожидает",
            "in_transit": "🚛 В пути",
            "delivered": "✅ Доставлено",
            "cancelled": "❌ Отменено"
        }
        status_text = status_mapping.get(status, "⏳ Ожидает")
        idx = self.status.findText(status_text)
        if idx >= 0:
            self.status.setCurrentIndex(idx)

        # Автомобиль
        car_id = shipment.get("car_id")
        if car_id:
            idx = self.car_combo.findData(car_id)
            if idx >= 0:
                self.car_combo.setCurrentIndex(idx)

        # Водитель
        driver_id = shipment.get("driver_id")
        if driver_id:
            idx = self.driver_combo.findData(driver_id)
            if idx >= 0:
                self.driver_combo.setCurrentIndex(idx)

        # Маршрут
        route_id = shipment.get("route_id")
        if route_id:
            idx = self.route_combo.findData(route_id)
            if idx >= 0:
                self.route_combo.setCurrentIndex(idx)

        # Тариф
        tariff_id = shipment.get("tariff_id")
        if tariff_id:
            idx = self.tariff_combo.findData(tariff_id)
            if idx >= 0:
                self.tariff_combo.setCurrentIndex(idx)

        # Расчетные поля
        self.calculate_cost()

    def check_car_capacity(self):
        """Проверить грузоподъемность автомобиля"""
        car_id = self.car_combo.currentData()
        weight = self.cargo_weight.value()

        if car_id and weight > 0:
            car = next((c for c in self.available_cars if c["id"] == car_id), None)
            if car:
                capacity_kg = car["load_capacity"] * 1000
                if weight > capacity_kg:
                    self.car_combo.setStyleSheet("QComboBox { border: 2px solid red; }")
                    return False
                else:
                    self.car_combo.setStyleSheet("")

        return True

    def calculate_cost(self):
        """Рассчитать стоимость перевозки"""
        route_id = self.route_combo.currentData()
        tariff_id = self.tariff_combo.currentData()

        if not route_id or not tariff_id:
            return

        # Находим маршрут и тариф
        route = next((r for r in self.available_routes if r["id"] == route_id), None)
        tariff = next((t for t in self.available_tariffs if t["id"] == tariff_id), None)

        if not route or not tariff:
            return

        # Обновляем информацию
        self.distance_label.setText(f"{route['distance_km']} км")
        self.price_per_km_label.setText(f"{tariff['price_per_km']} руб/км")
        self.min_price_label.setText(f"{tariff['min_price']} руб")

        # Рассчитываем стоимость
        distance = route['distance_km']
        price_per_km = tariff['price_per_km']
        min_price = tariff['min_price']

        cost = distance * price_per_km
        final_cost = max(cost, min_price)

        self.total_cost_label.setText(f"{final_cost:.2f} руб")

        # Подсвечиваем красным, если перегруз
        if not self.check_car_capacity():
            self.total_cost_label.setStyleSheet("color: red; font-weight: bold;")
            self.total_cost_label.setText(f"{final_cost:.2f} руб ⚠ ПЕРЕГРУЗ")
        else:
            self.total_cost_label.setStyleSheet("color: green; font-weight: bold;")

    def show_export_dialog(self):
        """Показать диалог выбора фильтра для экспорта перевозок"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Экспорт перевозок в Excel")
        dialog.setMinimumWidth(400)

        layout = QVBoxLayout(dialog)

        # Группа с вариантами экспорта
        group_box = QGroupBox("Выберите что экспортировать:")
        group_layout = QVBoxLayout()

        self.export_options_group = QButtonGroup(dialog)

        options = [
            ("Все перевозки", "all"),
            ("Текущие перевозки", "current"),
            ("Завершенные перевозки", "completed"),
            ("Отмененные перевозки", "cancelled"),
            ("Перевозки по статусу", "by_status")
        ]

        for text, option_id in options:
            radio = QRadioButton(text)
            radio.option_id = option_id
            self.export_options_group.addButton(radio)
            group_layout.addWidget(radio)

        # Выбираем первый вариант по умолчанию
        if options:
            self.export_options_group.buttons()[0].setChecked(True)

        group_box.setLayout(group_layout)
        layout.addWidget(group_box)

        # Выбор статуса (если выбрано "по статусу")
        self.status_combo = QComboBox()
        self.status_combo.setVisible(False)
        self.status_combo.addItems(["⏳ Ожидает", "🚛 В пути", "✅ Доставлено", "❌ Отменено"])

        def on_option_changed():
            selected_btn = self.export_options_group.checkedButton()
            if selected_btn and selected_btn.option_id == "by_status":
                self.status_combo.setVisible(True)
            else:
                self.status_combo.setVisible(False)

        self.export_options_group.buttonClicked.connect(on_option_changed)
        layout.addWidget(self.status_combo)

        # Кнопки
        buttons_layout = QHBoxLayout()
        export_btn = QPushButton("Экспорт")
        cancel_btn = QPushButton("Отмена")

        export_btn.clicked.connect(lambda: self.do_export(dialog))
        cancel_btn.clicked.connect(dialog.reject)

        buttons_layout.addWidget(export_btn)
        buttons_layout.addWidget(cancel_btn)
        layout.addLayout(buttons_layout)

        dialog.exec()

    def do_export(self, dialog):
        """Выполнить экспорт с выбранным фильтром"""
        # Получаем выбранный вариант
        selected_btn = self.export_options_group.checkedButton()
        if not selected_btn:
            QMessageBox.warning(dialog, "Ошибка", "Выберите вариант экспорта")
            return

        option_id = selected_btn.option_id

        # Дополнительные параметры
        extra_params = {}
        if option_id == "by_status":
            status_text = self.status_combo.currentText()
            # Преобразуем читаемый статус обратно
            status_mapping = {
                "⏳ Ожидает": "pending",
                "🚛 В пути": "in_transit",
                "✅ Доставлено": "delivered",
                "❌ Отменено": "cancelled"
            }
            extra_params["status"] = status_mapping.get(status_text, "pending")

        # Получаем сессию
        session = self.get_session()
        if not session:
            QMessageBox.warning(dialog, "Ошибка",
                                "Не удалось получить доступ к базе данных")
            return

        # Создаем диалог для выбора файла
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        file_name, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить файл Excel",
            f"перевозки_{option_id}_{timestamp}.xlsx",
            "Excel files (*.xlsx)"
        )

        if not file_name:
            return  # Пользователь отменил

        try:
            # Экспортируем данные с фильтром
            if self.export_shipments_to_excel(session, file_name, option_id, **extra_params):
                QMessageBox.information(
                    self,
                    "Экспорт завершен",
                    f"Данные успешно экспортированы в файл:\n{os.path.basename(file_name)}"
                )
                dialog.accept()
            else:
                QMessageBox.warning(dialog, "Ошибка",
                                    "Не удалось экспортировать данные")

        except Exception as e:
            QMessageBox.critical(dialog, "Ошибка экспорта",
                                 f"Произошла ошибка:\n{str(e)}")

    def get_session(self):
        """Получить сессию базы данных"""
        parent = self.parent()
        if parent and hasattr(parent, 'session'):
            return parent.session

        # Пробуем получить сессию из главного окна
        main_window = self.window()
        if main_window and hasattr(main_window, 'session'):
            return main_window.session

        return None

    def validate_and_accept(self):
        """Проверка данных и сохранение"""
        errors = []

        # Проверяем обязательные поля
        if not self.car_combo.currentData():
            errors.append("Выберите автомобиль")

        if not self.driver_combo.currentData():
            errors.append("Выберите водителя")

        if not self.route_combo.currentData():
            errors.append("Выберите маршрут")

        if not self.tariff_combo.currentData():
            errors.append("Выберите тариф")

        if self.cargo_weight.value() <= 0:
            errors.append("Вес груза должен быть больше 0")

        # Проверяем грузоподъемность
        if not self.check_car_capacity():
            errors.append("Вес груза превышает грузоподъемность автомобиля")

        if errors:
            QMessageBox.warning(self, "Ошибка", "\n".join(errors))
            return

        self.accept()

    def get_data(self):
        """Получить данные из формы"""
        qdt = self.shipment_date.dateTime()
        shipment_date = datetime.datetime(
            qdt.date().year(),
            qdt.date().month(),
            qdt.date().day(),
            qdt.time().hour(),
            qdt.time().minute()
        )

        # Преобразуем статус обратно в код
        status_mapping = {
            "⏳ Ожидает": "pending",
            "🚛 В пути": "in_transit",
            "✅ Доставлено": "delivered",
            "❌ Отменено": "cancelled"
        }
        status_code = status_mapping.get(self.status.currentText(), "pending")

        return {
            "shipment_date": shipment_date.isoformat(),
            "cargo_weight": self.cargo_weight.value(),
            "status": status_code,
            "car_id": self.car_combo.currentData(),
            "driver_id": self.driver_combo.currentData(),
            "route_id": self.route_combo.currentData(),
            "tariff_id": self.tariff_combo.currentData()
        }

    def export_shipments_to_excel(self, session, file_name, filter_type="all", **kwargs):
        """Экспорт перевозок в Excel с фильтрацией"""
        try:
            print(f"=== ДЕБАГ: Начало экспорта перевозок ===")
            print(f"1. Фильтр: {filter_type}")
            print(f"2. Доп. параметры: {kwargs}")
            print(f"3. Файл для сохранения: {file_name}")

            # Импортируем сервис перевозок
            try:
                from Services.Transportation.service import get_all_shipments, get_shipments_with_filters
                print("✅ 4. Импорт функций успешен")
            except ImportError as e:
                print(f"❌ 4. Ошибка импорта: {e}")
                QMessageBox.critical(self, "Ошибка", f"Не удалось импортировать модуль: {e}")
                return False

            import datetime

            print("5. Получаем данные из базы...")
            try:
                # Получаем данные в зависимости от фильтра
                if filter_type == "all":
                    shipments_data = get_all_shipments(session)
                    print(f"   Использован get_all_shipments")
                elif filter_type == "current":
                    shipments_data = get_shipments_with_filters(session, status=["pending", "in_transit"])
                    print(f"   Использован get_shipments_with_filters: status=['pending', 'in_transit']")
                elif filter_type == "completed":
                    shipments_data = get_shipments_with_filters(session, status=["delivered"])
                    print(f"   Использован get_shipments_with_filters: status=['delivered']")
                elif filter_type == "cancelled":
                    shipments_data = get_shipments_with_filters(session, status=["cancelled"])
                    print(f"   Использован get_shipments_with_filters: status=['cancelled']")
                elif filter_type == "by_status":
                    status = kwargs.get("status", "pending")
                    shipments_data = get_shipments_with_filters(session, status=[status])
                    print(f"   Использован get_shipments_with_filters: status=[{status}]")
                else:
                    shipments_data = get_all_shipments(session)
                    print(f"   Использован get_all_shipments (по умолчанию)")

                print(f"6. Получено перевозок: {len(shipments_data) if shipments_data else 0}")

            except Exception as e:
                print(f"❌ Ошибка при получении данных: {type(e).__name__}: {e}")
                import traceback
                traceback.print_exc()
                QMessageBox.critical(self, "Ошибка", f"Ошибка при получении данных:\n{str(e)}")
                return False

            if not shipments_data:
                print("7. ⚠ Нет данных после фильтрации")
                QMessageBox.warning(self, "Нет данных",
                                    f"Нет перевозок по выбранному фильтру: {filter_type}")
                return False

            # Проверяем структуру данных
            if shipments_data:
                print(f"8. Пример структуры данных первой перевозки:")
                print(f"   Ключи: {list(shipments_data[0].keys())}")
                print(f"   ID: {shipments_data[0].get('id')}")
                print(f"   Статус: {shipments_data[0].get('status')}")
                print(f"   Дата: {shipments_data[0].get('shipment_date')}")

            print("9. Создаем Excel файл...")
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                from openpyxl.utils import get_column_letter
                print("   ✅ Импорт openpyxl успешен")
            except ImportError as e:
                print(f"   ❌ Ошибка импорта openpyxl: {e}")
                QMessageBox.critical(self, "Ошибка", f"Не установлен openpyxl: {e}")
                return False

            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Перевозки"

                print("10. Настраиваем стили...")

                # Стили
                header_font = Font(bold=True, size=12, color="FFFFFF")
                header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                title_font = Font(bold=True, size=16, color="2C3E50")
                title_alignment = Alignment(horizontal="center", vertical="center")

                # Заголовок
                ws.merge_cells('A1:L1')
                title_cell = ws['A1']
                title_cell.value = "ОТЧЕТ ПО ПЕРЕВОЗКАМ"
                title_cell.font = title_font
                title_cell.alignment = title_alignment

                # Заголовки столбцов
                headers = [
                    ("ID", 8),
                    ("Дата", 18),
                    ("Тип груза", 15),
                    ("Вес (кг)", 12),
                    ("Статус", 12),
                    ("Автомобиль", 20),
                    ("Водитель", 20),
                    ("Маршрут", 25),
                    ("Расстояние (км)", 15),
                    ("Тариф (руб/км)", 15),
                    ("Мин. цена", 12),
                    ("Стоимость (руб)", 15)
                ]

                for col_idx, (header, width) in enumerate(headers, 1):
                    cell = ws.cell(row=3, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    ws.column_dimensions[get_column_letter(col_idx)].width = width

                print("11. Заполняем данные...")

                # Заполняем данными
                current_row = 4
                total_cost = 0
                total_weight = 0
                total_distance = 0

                # Маппинг статусов
                status_mapping = {
                    "pending": "⏳ Ожидает",
                    "in_transit": "🚛 В пути",
                    "delivered": "✅ Доставлено",
                    "cancelled": "❌ Отменено"
                }

                for i, shipment in enumerate(shipments_data):
                    # Отладочный вывод для первых 2 записей
                    if i < 2:
                        print(f"   Обработка записи {i + 1}: ID={shipment.get('id')}")

                    # Форматируем дату
                    shipment_date = shipment.get("shipment_date", "")
                    if shipment_date:
                        try:
                            dt = datetime.datetime.fromisoformat(shipment_date)
                            date_str = dt.strftime("%d.%m.%Y %H:%M")
                        except Exception as e:
                            print(f"   ❌ Ошибка форматирования даты: {e}")
                            date_str = str(shipment_date)
                    else:
                        date_str = ""

                    # Получаем информацию
                    car_info = shipment.get("car_info", {})
                    driver_info = shipment.get("driver_info", {})
                    route_info = shipment.get("route_info", {})
                    tariff_info = shipment.get("tariff_info", {})

                    # Проверяем наличие данных
                    car_text = f"{car_info.get('brand', '')} ({car_info.get('license_plate', '')})"
                    driver_text = f"{driver_info.get('full_name', '')}"
                    if driver_info.get('license_number'):
                        driver_text += f" ({driver_info.get('license_number')})"

                    route_text = ""
                    if route_info.get('origin') and route_info.get('destination'):
                        route_text = f"{route_info['origin']} → {route_info['destination']}"

                    # Статус с иконкой
                    status = shipment.get("status", "pending")
                    status_text = status_mapping.get(status, "⏳ Ожидает")

                    # Стоимость
                    cost = shipment.get("total_cost", 0)

                    # Записываем данные
                    ws.cell(row=current_row, column=1, value=str(shipment.get("id", "")))
                    ws.cell(row=current_row, column=2, value=date_str)
                    ws.cell(row=current_row, column=4, value=shipment.get("cargo_weight", 0))
                    ws.cell(row=current_row, column=5, value=status_text)
                    ws.cell(row=current_row, column=6, value=car_text)
                    ws.cell(row=current_row, column=7, value=driver_text)
                    ws.cell(row=current_row, column=8, value=route_text)
                    ws.cell(row=current_row, column=9, value=route_info.get('distance_km', 0))
                    ws.cell(row=current_row, column=10, value=tariff_info.get('price_per_km', 0))
                    ws.cell(row=current_row, column=11, value=tariff_info.get('min_price', 0))
                    ws.cell(row=current_row, column=12, value=cost)

                    # Форматирование числовых полей
                    ws.cell(row=current_row, column=4).number_format = '#,##0.0" кг"'
                    ws.cell(row=current_row, column=9).number_format = '#,##0" км"'
                    ws.cell(row=current_row, column=10).number_format = '#,##0.00" руб"'
                    ws.cell(row=current_row, column=11).number_format = '#,##0" руб"'
                    ws.cell(row=current_row, column=12).number_format = '#,##0.00" руб"'

                    # Собираем статистику
                    total_cost += cost
                    total_weight += shipment.get("cargo_weight", 0)
                    total_distance += route_info.get('distance_km', 0)

                    # Подсветка строк по статусу
                    if status == "delivered":
                        for col in range(1, 13):
                            ws.cell(row=current_row, column=col).fill = PatternFill(
                                start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"
                            )
                    elif status == "cancelled":
                        for col in range(1, 13):
                            ws.cell(row=current_row, column=col).fill = PatternFill(
                                start_color="FADBD8", end_color="FADBD8", fill_type="solid"
                            )

                    current_row += 1

                print(f"12. Заполнено строк: {current_row - 4}")

                # Добавляем границы для данных
                thin_border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))

                for row in ws.iter_rows(min_row=3, max_row=current_row - 1, max_col=12):
                    for cell in row:
                        cell.border = thin_border

                print("13. Добавляем статистику...")

                # Добавляем статистику
                stats_row = current_row + 2

                ws.merge_cells(f'A{stats_row}:L{stats_row}')
                ws.cell(row=stats_row, column=1, value="СТАТИСТИКА").font = Font(bold=True, size=14)
                ws.cell(row=stats_row, column=1).alignment = Alignment(horizontal="center")

                # Информация о фильтре
                filter_names = {
                    "all": "Все перевозки",
                    "current": "Текущие перевозки (ожидает/в пути)",
                    "completed": "Завершенные перевозки",
                    "cancelled": "Отмененные перевозки",
                    "by_status": f"Перевозки по статусу: {status_mapping.get(kwargs.get('status', 'pending'), 'Ожидает')}"
                }

                ws.cell(row=stats_row + 1, column=1, value=f"Фильтр: {filter_names.get(filter_type, 'Все перевозки')}")
                ws.cell(row=stats_row + 1, column=2, value=f"Количество перевозок: {len(shipments_data)}")

                if shipments_data:
                    # Подсчет по статусам
                    status_counts = {}
                    for shipment in shipments_data:
                        status = shipment.get("status", "pending")
                        status_counts[status] = status_counts.get(status, 0) + 1

                    ws.cell(row=stats_row + 2, column=1, value=f"Ожидает: {status_counts.get('pending', 0)}")
                    ws.cell(row=stats_row + 2, column=2, value=f"В пути: {status_counts.get('in_transit', 0)}")
                    ws.cell(row=stats_row + 2, column=3, value=f"Доставлено: {status_counts.get('delivered', 0)}")
                    ws.cell(row=stats_row + 2, column=4, value=f"Отменено: {status_counts.get('cancelled', 0)}")

                    ws.cell(row=stats_row + 3, column=1, value=f"Общий вес: {total_weight:.1f} кг")
                    ws.cell(row=stats_row + 3, column=2, value=f"Общая дистанция: {total_distance:.0f} км")
                    ws.cell(row=stats_row + 3, column=3, value=f"Общая стоимость: {total_cost:.2f} руб")

                    if total_weight > 0:
                        avg_cost_per_kg = total_cost / total_weight
                        ws.cell(row=stats_row + 4, column=1,
                                value=f"Средняя стоимость за кг: {avg_cost_per_kg:.2f} руб/кг")

                    if total_distance > 0:
                        avg_cost_per_km = total_cost / total_distance
                        ws.cell(row=stats_row + 4, column=2,
                                value=f"Средняя стоимость за км: {avg_cost_per_km:.2f} руб/км")

                # Добавляем дату экспорта
                ws.cell(row=stats_row + 6, column=1,
                        value=f"Экспортировано: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")

                # Автоподбор ширины для некоторых столбцов
                for col in [6, 7, 8]:  # Автомобиль, Водитель, Маршрут
                    ws.column_dimensions[get_column_letter(col)].auto_size = True

                print(f"14. Сохраняем файл: {file_name}")
                wb.save(file_name)
                print("✅ 15. Файл успешно сохранен!")

                return True

            except Exception as e:
                print(f"❌ Ошибка при создании Excel: {type(e).__name__}: {e}")
                import traceback
                traceback.print_exc()
                raise  # Пробрасываем дальше

        except Exception as e:
            print(f"❌ Критическая ошибка в export_shipments_to_excel: {type(e).__name__}: {e}")
            import traceback
            traceback.print_exc()  # Выводим полный стек вызовов

            # Показываем пользователю понятное сообщение
            error_msg = str(e)
            if "Permission denied" in error_msg:
                error_msg = "Нет прав для записи в выбранную папку. Выберите другую папку."
            elif "directory" in error_msg.lower():
                error_msg = "Указанная папка не существует. Проверьте путь."
            elif "openpyxl" in error_msg.lower():
                error_msg = "Ошибка работы с Excel файлом. Убедитесь, что файл не открыт в другой программе."

            QMessageBox.critical(self, "Ошибка экспорта",
                                 f"Произошла ошибка:\n{error_msg}")
            return False