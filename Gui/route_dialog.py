# gui/route_dialog.py
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QLabel,
    QLineEdit, QPushButton, QDoubleSpinBox,
    QHBoxLayout, QRadioButton, QButtonGroup,
    QGroupBox, QMessageBox, QFileDialog
)
from datetime import datetime
import os


class CreateRouteDialog(QDialog):
    def __init__(self, parent=None, route=None):
        super().__init__(parent)
        self.setWindowTitle("Редактирование маршрута" if route else "Создать маршрут")
        self.route = route

        # Поля формы
        self.origin = QLineEdit()
        self.origin.setPlaceholderText("Например: Москва")

        self.destination = QLineEdit()
        self.destination.setPlaceholderText("Например: Санкт-Петербург")

        self.distance = QDoubleSpinBox()
        self.distance.setMaximum(100000)
        self.distance.setMinimum(1)
        self.distance.setSuffix(" км")
        self.distance.setDecimals(1)
        self.distance.setSingleStep(10)

        self.avg_time = QDoubleSpinBox()
        self.avg_time.setMaximum(1000)
        self.avg_time.setMinimum(0.1)
        self.avg_time.setSuffix(" ч")
        self.avg_time.setDecimals(1)
        self.avg_time.setSingleStep(0.5)

        self.road_type = QLineEdit()
        self.road_type.setPlaceholderText("Например: магистраль, грунтовая")

        # Кнопка сохранения
        btn_text = "Сохранить изменения" if route else "Создать маршрут"
        self.save_btn = QPushButton(btn_text)
        self.save_btn.clicked.connect(self.validate_and_accept)

        # Кнопка экспорта (НОВАЯ)
        self.export_btn = QPushButton("Экспорт маршрутов в Excel")
        self.export_btn.clicked.connect(self.show_export_dialog)

        # Кнопка отмены
        self.cancel_btn = QPushButton("Отмена")
        self.cancel_btn.clicked.connect(self.reject)

        # Разметка
        layout = QVBoxLayout(self)

        layout.addWidget(QLabel("Пункт отправления *"))
        layout.addWidget(self.origin)

        layout.addWidget(QLabel("Пункт назначения *"))
        layout.addWidget(self.destination)

        layout.addWidget(QLabel("Расстояние *"))
        layout.addWidget(self.distance)

        layout.addWidget(QLabel("Среднее время *"))
        layout.addWidget(self.avg_time)

        layout.addWidget(QLabel("Тип дороги"))
        layout.addWidget(self.road_type)

        # Горизонтальный layout для кнопок
        buttons_layout = QHBoxLayout()
        buttons_layout.addWidget(self.save_btn)
        buttons_layout.addWidget(self.export_btn)
        buttons_layout.addWidget(self.cancel_btn)
        layout.addLayout(buttons_layout)

        # Если передан маршрут для редактирования - заполняем поля
        if route:
            self.load_data(route)

    def load_data(self, route):
        """Заполнить поля данными маршрута"""
        self.origin.setText(route.get("origin", ""))
        self.destination.setText(route.get("destination", ""))
        self.distance.setValue(route.get("distance_km", 0))
        self.avg_time.setValue(route.get("avg_time_hours", 0))
        self.road_type.setText(route.get("road_type", ""))

    def validate_and_accept(self):
        """Проверка данных перед сохранением"""
        origin = self.origin.text().strip()
        destination = self.destination.text().strip()
        road_type = self.road_type.text().strip()

        errors = []

        if not origin:
            errors.append("Укажите пункт отправления")
        if not destination:
            errors.append("Укажите пункт назначения")
        if origin and destination and origin.lower() == destination.lower():
            errors.append("Пункты отправления и назначения не могут совпадать")

        if self.distance.value() <= 0:
            errors.append("Расстояние должно быть больше 0")

        if self.avg_time.value() <= 0:
            errors.append("Среднее время должно быть больше 0")

        if errors:
            QMessageBox.warning(self, "Ошибка заполнения", "\n".join(errors))
            return

        self.accept()

    def get_data(self):
        """Получить данные из формы"""
        return {
            "origin": self.origin.text().strip(),
            "destination": self.destination.text().strip(),
            "distance_km": self.distance.value(),
            "avg_time_hours": self.avg_time.value(),
            "road_type": self.road_type.text().strip(),
        }

    def show_export_dialog(self):
        """Показать диалог выбора фильтра для экспорта маршрутов"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Экспорт маршрутов в Excel")
        dialog.setMinimumWidth(350)

        layout = QVBoxLayout(dialog)

        # Группа с вариантами экспорта
        group_box = QGroupBox("Выберите что экспортировать:")
        group_layout = QVBoxLayout()

        self.export_options_group = QButtonGroup(dialog)

        options = [
            ("Все маршруты", "all"),
            ("Длинные маршруты (> 500 км)", "long"),
            ("Короткие маршруты (< 200 км)", "short"),
            ("Магистральные маршруты", "highway"),
            ("Городские маршруты", "city")
        ]

        for text, option_id in options:
            radio = QRadioButton(text)
            radio.option_id = option_id
            self.export_options_group.addButton(radio)
            group_layout.addWidget(radio)

        # Выбираем первый вариант по умолчанию
        if options:
            self.export_options_group.buttons()[0].setChecked(True)

        group_box.setLayout(group_layout)
        layout.addWidget(group_box)

        # Кнопки
        buttons_layout = QHBoxLayout()
        export_btn = QPushButton("Экспорт")
        cancel_btn = QPushButton("Отмена")

        export_btn.clicked.connect(lambda: self.do_export(dialog))
        cancel_btn.clicked.connect(dialog.reject)

        buttons_layout.addWidget(export_btn)
        buttons_layout.addWidget(cancel_btn)
        layout.addLayout(buttons_layout)

        dialog.exec()

    def do_export(self, dialog):
        """Выполнить экспорт с выбранным фильтром"""
        # Получаем выбранный вариант
        selected_btn = self.export_options_group.checkedButton()
        if not selected_btn:
            QMessageBox.warning(dialog, "Ошибка", "Выберите вариант экспорта")
            return

        option_id = selected_btn.option_id

        # Получаем сессию
        session = self.get_session()
        if not session:
            QMessageBox.warning(dialog, "Ошибка",
                                "Не удалось получить доступ к базе данных")
            return

        # Создаем диалог для выбора файла
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        file_name, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить файл Excel",
            f"маршруты_{option_id}_{timestamp}.xlsx",
            "Excel files (*.xlsx)"
        )

        if not file_name:
            return  # Пользователь отменил

        try:
            # Экспортируем данные с фильтром
            if self.export_routes_to_excel(session, file_name, option_id):
                QMessageBox.information(
                    self,
                    "Экспорт завершен",
                    f"Данные успешно экспортированы в файл:\n{os.path.basename(file_name)}"
                )
                dialog.accept()
            else:
                QMessageBox.warning(dialog, "Ошибка",
                                    "Не удалось экспортировать данные")

        except Exception as e:
            QMessageBox.critical(dialog, "Ошибка экспорта",
                                 f"Произошла ошибка:\n{str(e)}")

    def get_session(self):
        """Получить сессию базы данных"""
        parent = self.parent()
        if parent and hasattr(parent, 'session'):
            return parent.session

        # Пробуем получить сессию из главного окна
        main_window = self.window()
        if main_window and hasattr(main_window, 'session'):
            return main_window.session

        return None

    def export_routes_to_excel(self, session, file_name, filter_type="all"):
        """Экспорт маршрутов в Excel с фильтрацией"""
        try:
            # Импортируем сервис маршрутов
            from Services.Route.services import get_all_routes, get_routes_with_filters

            # Получаем данные в зависимости от фильтра
            if filter_type == "all":
                routes_data = get_all_routes(session)
            elif filter_type == "long":
                routes_data = get_routes_with_filters(session, min_distance=500)
            elif filter_type == "short":
                routes_data = get_routes_with_filters(session, max_distance=200)
            elif filter_type == "highway":
                routes_data = get_routes_with_filters(session, road_type="магистраль")
            elif filter_type == "city":
                routes_data = get_routes_with_filters(session, road_type="город")
            else:
                routes_data = get_all_routes(session)

            if not routes_data:
                QMessageBox.warning(self, "Нет данных",
                                    f"Нет маршрутов по выбранному фильтру: {filter_type}")
                return False

            # Создаем Excel файл
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Маршруты"

            # Стили для заголовков
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="D4E6F1", end_color="D4E6F1", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Заголовки столбцов
            headers = [
                "ID", "Откуда", "Куда",
                "Расстояние (км)", "Среднее время (ч)",
                "Тип дороги", "Скорость (км/ч)"
            ]

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Заполняем данными
            total_distance = 0
            for row_idx, route in enumerate(routes_data, 2):
                # Конвертируем ID в строку
                route_id = route.get("id", "")
                if hasattr(route_id, '__str__'):
                    route_id = str(route_id)

                # Рассчитываем среднюю скорость
                distance = route.get("distance_km", 0)
                time = route.get("avg_time_hours", 1)
                avg_speed = distance / time if time > 0 else 0

                total_distance += distance

                ws.cell(row=row_idx, column=1, value=route_id)
                ws.cell(row=row_idx, column=2, value=route.get("origin", ""))
                ws.cell(row=row_idx, column=3, value=route.get("destination", ""))
                ws.cell(row=row_idx, column=4, value=distance)
                ws.cell(row=row_idx, column=5, value=time)
                ws.cell(row=row_idx, column=6, value=route.get("road_type", ""))
                ws.cell(row=row_idx, column=7, value=round(avg_speed, 1))

            # Автоподбор ширины столбцов
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        cell_value = str(cell.value) if cell.value is not None else ""
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Добавляем статистику
            last_row = ws.max_row + 2
            filter_names = {
                "all": "Все маршруты",
                "long": "Длинные маршруты (> 500 км)",
                "short": "Короткие маршруты (< 200 км)",
                "highway": "Магистральные маршруты",
                "city": "Городские маршруты"
            }

            ws.cell(row=last_row, column=1, value=f"Фильтр: {filter_names.get(filter_type, 'Все маршруты')}")
            ws.cell(row=last_row, column=2, value=f"Количество маршрутов: {len(routes_data)}")
            ws.cell(row=last_row, column=3, value=f"Общее расстояние: {total_distance} км")

            # Рассчитываем среднюю скорость по всем маршрутам
            avg_speed_all = 0
            if routes_data:
                total_time = sum(r.get("avg_time_hours", 0) for r in routes_data)
                if total_time > 0:
                    avg_speed_all = total_distance / total_time

            ws.cell(row=last_row, column=4, value=f"Средняя скорость: {round(avg_speed_all, 1)} км/ч")

            # Добавляем дату экспорта
            ws.cell(row=last_row + 1, column=1,
                    value=f"Экспортировано: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

            # Сохраняем файл
            wb.save(file_name)
            return True

        except ImportError as e:
            QMessageBox.critical(self, "Ошибка",
                                 f"Не удалось импортировать модули: {str(e)}")
            return False
        except Exception as e:
            print(f"Ошибка при экспорте маршрутов: {str(e)}")
            return False