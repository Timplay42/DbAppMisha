# Gui/car_dialog.py
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QLabel,
    QLineEdit, QPushButton, QDoubleSpinBox,
    QMessageBox, QHBoxLayout, QRadioButton,
    QButtonGroup, QGroupBox, QFileDialog
)
from PySide6.QtCore import Qt
from datetime import datetime
import os


class CarDialog(QDialog):

    def __init__(self, parent=None, car=None):
        super().__init__(parent)
        self.setWindowTitle("Редактирование машины" if car else "Добавление машины")
        self.setMinimumWidth(400)

        self.car = car

        self.brand_input = QLineEdit()
        self.brand_input.setPlaceholderText("Например: Volvo")

        self.license_plate_input = QLineEdit()
        self.license_plate_input.setPlaceholderText("Например: A123BC123")

        self.load_capacity_input = QDoubleSpinBox()
        self.load_capacity_input.setSuffix(" т")
        self.load_capacity_input.setMinimum(0.1)
        self.load_capacity_input.setMaximum(100)
        self.load_capacity_input.setDecimals(1)
        self.load_capacity_input.setSingleStep(0.5)

        self.body_type_input = QLineEdit()
        self.body_type_input.setPlaceholderText("Например: Бортовой, Рефрижератор")

        self.fuel_consumption_input = QDoubleSpinBox()
        self.fuel_consumption_input.setSuffix(" л/100км")
        self.fuel_consumption_input.setMinimum(5)
        self.fuel_consumption_input.setMaximum(100)
        self.fuel_consumption_input.setDecimals(1)
        self.fuel_consumption_input.setSingleStep(0.5)

        # Кнопка сохранения
        btn_text = "Сохранить изменения" if car else "Добавить машину"
        self.save_btn = QPushButton(btn_text)
        self.save_btn.clicked.connect(self.validate_and_accept)

        # Кнопка экспорта (НОВАЯ КНОПКА)
        self.export_btn = QPushButton("Экспорт машин в Excel")
        self.export_btn.clicked.connect(self.show_export_dialog)

        # Кнопка отмены
        self.cancel_btn = QPushButton("Отмена")
        self.cancel_btn.clicked.connect(self.reject)

        # Разметка
        layout = QVBoxLayout(self)

        # Марка
        layout.addWidget(QLabel("Марка автомобиля *"))
        layout.addWidget(self.brand_input)

        # Госномер
        layout.addWidget(QLabel("Государственный номер *"))
        layout.addWidget(self.license_plate_input)

        # Грузоподъемность
        layout.addWidget(QLabel("Грузоподъемность *"))
        layout.addWidget(self.load_capacity_input)

        # Тип кузова
        layout.addWidget(QLabel("Тип кузова *"))
        layout.addWidget(self.body_type_input)

        # Расход топлива
        layout.addWidget(QLabel("Расход топлива *"))
        layout.addWidget(self.fuel_consumption_input)

        # Кнопки в горизонтальном layout для лучшего вида
        buttons_layout = QHBoxLayout()
        buttons_layout.addWidget(self.save_btn)
        buttons_layout.addWidget(self.export_btn)
        buttons_layout.addWidget(self.cancel_btn)

        layout.addLayout(buttons_layout)

        # Если передан объект машины - заполняем поля
        if car:
            self.load_data(car)

    def load_data(self, car):
        """Заполнить поля данными машины"""
        self.brand_input.setText(car.get("brand", ""))
        self.license_plate_input.setText(car.get("license_plate", ""))
        self.load_capacity_input.setValue(car.get("load_capacity", 0))
        self.body_type_input.setText(car.get("body_type", ""))
        self.fuel_consumption_input.setValue(car.get("fuel_consumption", 0))

    def validate_and_accept(self):
        """Проверка данных перед сохранением"""
        brand = self.brand_input.text().strip()
        license_plate = self.license_plate_input.text().strip()
        body_type = self.body_type_input.text().strip()

        errors = []

        if not brand:
            errors.append("Укажите марку автомобиля")
        if not license_plate:
            errors.append("Укажите государственный номер")
        if not body_type:
            errors.append("Укажите тип кузова")

        if self.load_capacity_input.value() <= 0:
            errors.append("Грузоподъемность должна быть больше 0")

        if self.fuel_consumption_input.value() <= 0:
            errors.append("Расход топлива должен быть больше 0")

        if errors:
            QMessageBox.warning(self, "Ошибка заполнения", "\n".join(errors))
            return

        self.accept()

    def get_data(self):
        """Получить данные из формы"""
        return {
            "brand": self.brand_input.text().strip(),
            "license_plate": self.license_plate_input.text().strip(),
            "load_capacity": self.load_capacity_input.value(),
            "body_type": self.body_type_input.text().strip(),
            "fuel_consumption": self.fuel_consumption_input.value()
        }

    def show_export_dialog(self):
        """Показать диалог выбора фильтра для экспорта"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Экспорт машин в Excel")
        dialog.setMinimumWidth(350)

        layout = QVBoxLayout(dialog)

        # Группа с вариантами экспорта
        group_box = QGroupBox("Выберите что экспортировать:")
        group_layout = QVBoxLayout()

        self.export_options_group = QButtonGroup(dialog)

        options = [
            ("Все машины", "all"),
            ("Свободные машины", "free"),
            ("Машины с грузоподъемностью > 10т", "heavy"),
            ("Экономичные машины (< 15 л/100км)", "efficient"),
            ("Машины с водителями", "with_driver")
        ]

        for text, option_id in options:
            radio = QRadioButton(text)
            radio.option_id = option_id
            self.export_options_group.addButton(radio)
            group_layout.addWidget(radio)

        # Выбираем первый вариант по умолчанию
        if options:
            self.export_options_group.buttons()[0].setChecked(True)

        group_box.setLayout(group_layout)
        layout.addWidget(group_box)

        # Кнопки
        buttons_layout = QHBoxLayout()
        export_btn = QPushButton("Экспорт")
        cancel_btn = QPushButton("Отмена")

        export_btn.clicked.connect(lambda: self.do_export(dialog))
        cancel_btn.clicked.connect(dialog.reject)

        buttons_layout.addWidget(export_btn)
        buttons_layout.addWidget(cancel_btn)
        layout.addLayout(buttons_layout)

        dialog.exec()

    def do_export(self, dialog):
        """Выполнить экспорт с выбранным фильтром"""
        # Получаем выбранный вариант
        selected_btn = self.export_options_group.checkedButton()
        if not selected_btn:
            QMessageBox.warning(dialog, "Ошибка", "Выберите вариант экспорта")
            return

        option_id = selected_btn.option_id

        # Получаем сессию
        session = self.get_session()
        if not session:
            QMessageBox.warning(dialog, "Ошибка",
                                "Не удалось получить доступ к базе данных")
            return

        # Создаем диалог для выбора файла
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        file_name, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить файл Excel",
            f"машины_{option_id}_{timestamp}.xlsx",
            "Excel files (*.xlsx)"
        )

        if not file_name:
            return  # Пользователь отменил

        try:
            # Экспортируем данные с фильтром
            if self.export_cars_to_excel(session, file_name, option_id):
                QMessageBox.information(
                    self,
                    "Экспорт завершен",
                    f"Данные успешно экспортированы в файл:\n{os.path.basename(file_name)}"
                )
                dialog.accept()
            else:
                QMessageBox.warning(dialog, "Ошибка",
                                    "Не удалось экспортировать данные")

        except Exception as e:
            QMessageBox.critical(dialog, "Ошибка экспорта",
                                 f"Произошла ошибка:\n{str(e)}")

    def get_session(self):
        """Получить сессию базы данных"""
        parent = self.parent()
        if parent and hasattr(parent, 'session'):
            return parent.session

        # Пробуем получить сессию из главного окна
        main_window = self.window()
        if main_window and hasattr(main_window, 'session'):
            return main_window.session

        return None

    def export_cars_to_excel(self, session, file_name, filter_type="all"):
        """Экспорт машин в Excel с фильтрацией"""
        try:
            # Получаем данные в зависимости от фильтра
            from Services.Car.services import (
                get_all_cars_with_drivers, get_free_cars,
                get_cars_by_load_capacity, get_cars_by_fuel_consumption
            )

            if filter_type == "all":
                cars_data = get_all_cars_with_drivers(session)
            elif filter_type == "free":
                cars_data = get_free_cars(session)
            elif filter_type == "heavy":
                cars_data = get_cars_by_load_capacity(session, 10.0)
            elif filter_type == "efficient":
                cars_data = get_cars_by_fuel_consumption(session, 15.0)
            elif filter_type == "with_driver":
                cars_data = get_all_cars_with_drivers(session)
                cars_data = [car for car in cars_data if car.get("has_driver", False)]
            else:
                cars_data = get_all_cars_with_drivers(session)

            if not cars_data:
                QMessageBox.warning(self, "Нет данных",
                                    f"Нет машин по выбранному фильтру: {filter_type}")
                return False

            # Создаем Excel файл
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Машины"

            # Стили для заголовков
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Заголовки столбцов
            headers = [
                "ID", "Марка", "Госномер",
                "Грузоподъемность (т)", "Тип кузова",
                "Расход топлива (л/100км)", "Статус", "Водитель"
            ]

            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Заполняем данными
            for row_idx, car in enumerate(cars_data, 2):
                # Конвертируем ID в строку, если это UUID
                car_id = car.get("id", "")
                if hasattr(car_id, '__str__'):
                    car_id = str(car_id)

                ws.cell(row=row_idx, column=1, value=car_id)
                ws.cell(row=row_idx, column=2, value=car.get("brand", ""))
                ws.cell(row=row_idx, column=3, value=car.get("license_plate", ""))
                ws.cell(row=row_idx, column=4, value=car.get("load_capacity", 0))
                ws.cell(row=row_idx, column=5, value=car.get("body_type", ""))
                ws.cell(row=row_idx, column=6, value=car.get("fuel_consumption", 0))
                ws.cell(row=row_idx, column=7, value=car.get("status", "Свободна"))
                ws.cell(row=row_idx, column=8, value=car.get("driver_name", ""))

            # Автоподбор ширины столбцов
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        cell_value = str(cell.value) if cell.value is not None else ""
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Добавляем информацию о фильтре
            last_row = ws.max_row + 2
            filter_names = {
                "all": "Все машины",
                "free": "Свободные машины",
                "heavy": "Машины с грузоподъемностью > 10т",
                "efficient": "Экономичные машины (< 15 л/100км)",
                "with_driver": "Машины с водителями"
            }

            ws.cell(row=last_row, column=1, value=f"Фильтр: {filter_names.get(filter_type, 'Все машины')}")
            ws.cell(row=last_row, column=2, value=f"Всего машин: {len(cars_data)}")

            if filter_type == "all":
                free_count = sum(1 for c in cars_data if c.get("status") == "Свободна")
                busy_count = sum(1 for c in cars_data if c.get("status") == "Занята")
                ws.cell(row=last_row, column=3, value=f"Свободных: {free_count}")
                ws.cell(row=last_row, column=4, value=f"Занятых: {busy_count}")

            # Добавляем дату экспорта
            ws.cell(row=last_row + 1, column=1,
                    value=f"Экспортировано: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

            # Сохраняем файл
            wb.save(file_name)
            return True

        except Exception as e:
            print(f"Ошибка при экспорте: {str(e)}")
            return False