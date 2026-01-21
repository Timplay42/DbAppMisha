# Gui/tariff_dialog.py
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel,
    QLineEdit, QPushButton, QDoubleSpinBox, QDateTimeEdit,
    QMessageBox, QComboBox, QGroupBox, QGridLayout,
    QTextEdit, QCheckBox, QRadioButton, QButtonGroup,
    QFileDialog
)
from PySide6.QtCore import QDateTime, Qt
from PySide6.QtGui import QFont
import datetime
import os


class TariffDialog(QDialog):
    """Диалог для создания/редактирования тарифа"""

    def __init__(self, parent=None, tariff=None, cargo_types=None):
        super().__init__(parent)
        self.setWindowTitle("Редактирование тарифа" if tariff else "Создание тарифа")
        self.setMinimumWidth(600)

        self.tariff = tariff
        self.cargo_types = cargo_types or []

        # Виджеты
        self.cargo_type_combo = QComboBox()
        self.cargo_type_combo.setEditable(True)  # Разрешаем ввод новых типов
        if self.cargo_types:
            self.cargo_type_combo.addItems(self.cargo_types)

        self.price_per_km_input = QDoubleSpinBox()
        self.price_per_km_input.setRange(0.01, 1000)
        self.price_per_km_input.setSuffix(" руб/км")
        self.price_per_km_input.setDecimals(2)
        self.price_per_km_input.setSingleStep(0.1)

        self.min_price_input = QDoubleSpinBox()
        self.min_price_input.setRange(0, 100000)
        self.min_price_input.setSuffix(" руб")
        self.min_price_input.setDecimals(2)
        self.min_price_input.setSingleStep(100)

        self.date_start_input = QDateTimeEdit()
        self.date_start_input.setCalendarPopup(True)
        self.date_start_input.setDateTime(QDateTime.currentDateTime())
        self.date_start_input.setDisplayFormat("dd.MM.yyyy HH:mm")

        self.date_end_input = QDateTimeEdit()
        self.date_end_input.setCalendarPopup(True)
        self.date_end_input.setDateTime(QDateTime.currentDateTime().addYears(1))
        self.date_end_input.setDisplayFormat("dd.MM.yyyy HH:mm")

        self.no_end_date_check = QCheckBox("Бессрочный тариф")
        self.no_end_date_check.stateChanged.connect(self.toggle_end_date)

        self.description_input = QTextEdit()
        self.description_input.setMaximumHeight(100)
        self.description_input.setPlaceholderText("Описание тарифа...")

        # Кнопки
        self.save_btn = QPushButton("💾 Сохранить")
        self.save_btn.clicked.connect(self.validate_and_accept)

        self.export_btn = QPushButton("📊 Экспорт тарифов")
        self.export_btn.clicked.connect(self.show_export_dialog)

        self.cancel_btn = QPushButton("❌ Отмена")
        self.cancel_btn.clicked.connect(self.reject)

        self.calculate_btn = QPushButton("📈 Пример расчета")
        self.calculate_btn.clicked.connect(self.show_calculation_example)

        self.setup_ui()

        if tariff:
            self.load_data(tariff)

    def setup_ui(self):
        """Настройка интерфейса"""
        layout = QVBoxLayout(self)

        # Основная информация
        info_group = QGroupBox("Основная информация")
        info_layout = QGridLayout()

        info_layout.addWidget(QLabel("Тип груза *"), 0, 0)
        info_layout.addWidget(self.cargo_type_combo, 0, 1)

        info_layout.addWidget(QLabel("Цена за километр *"), 1, 0)
        info_layout.addWidget(self.price_per_km_input, 1, 1)

        info_layout.addWidget(QLabel("Минимальная цена *"), 2, 0)
        info_layout.addWidget(self.min_price_input, 2, 1)

        info_layout.addWidget(QLabel("Дата начала *"), 3, 0)
        info_layout.addWidget(self.date_start_input, 3, 1)

        info_layout.addWidget(QLabel("Дата окончания"), 4, 0)
        date_layout = QHBoxLayout()
        date_layout.addWidget(self.date_end_input)
        date_layout.addWidget(self.no_end_date_check)
        info_layout.addLayout(date_layout, 4, 1)

        info_group.setLayout(info_layout)
        layout.addWidget(info_group)

        # Описание
        desc_group = QGroupBox("Описание (необязательно)")
        desc_layout = QVBoxLayout()
        desc_layout.addWidget(self.description_input)
        desc_group.setLayout(desc_layout)
        layout.addWidget(desc_group)

        # Кнопки
        btn_layout = QHBoxLayout()
        btn_layout.addWidget(self.calculate_btn)
        btn_layout.addWidget(self.export_btn)
        btn_layout.addStretch()
        btn_layout.addWidget(self.save_btn)
        btn_layout.addWidget(self.cancel_btn)

        layout.addLayout(btn_layout)

    def toggle_end_date(self, state):
        """Включить/выключить поле даты окончания"""
        self.date_end_input.setEnabled(state == Qt.Unchecked)

    def load_data(self, tariff):
        """Загрузить данные тарифа"""
        # Тип груза
        cargo_type = tariff.get("cargo_type", "")
        idx = self.cargo_type_combo.findText(cargo_type)
        if idx >= 0:
            self.cargo_type_combo.setCurrentIndex(idx)
        else:
            self.cargo_type_combo.setCurrentText(cargo_type)

        # Цены
        self.price_per_km_input.setValue(tariff.get("price_per_km", 0))
        self.min_price_input.setValue(tariff.get("min_price", 0))

        # Даты
        date_start = tariff.get("date_start")
        if date_start:
            if isinstance(date_start, str):
                date_start = datetime.datetime.fromisoformat(date_start)
            qdt = QDateTime(
                QDateTime.fromString(date_start.isoformat(), Qt.ISODate)
            )
            self.date_start_input.setDateTime(qdt)

        date_end = tariff.get("date_end")
        if date_end:
            if isinstance(date_end, str):
                date_end = datetime.datetime.fromisoformat(date_end)
            qdt = QDateTime(
                QDateTime.fromString(date_end.isoformat(), Qt.ISODate)
            )
            self.date_end_input.setDateTime(qdt)
            self.no_end_date_check.setChecked(False)
        else:
            self.no_end_date_check.setChecked(True)
            self.date_end_input.setEnabled(False)

        # Описание
        description = tariff.get("description", "")
        self.description_input.setText(description)

    def show_calculation_example(self):
        """Показать пример расчета стоимости"""
        price_per_km = self.price_per_km_input.value()
        min_price = self.min_price_input.value()

        examples = [
            {"distance": 50, "description": "Короткая перевозка"},
            {"distance": 200, "description": "Средняя перевозка"},
            {"distance": 500, "description": "Дальняя перевозка"},
        ]

        result_text = "Пример расчета стоимости:\n\n"
        for example in examples:
            cost = example["distance"] * price_per_km
            final_cost = max(cost, min_price)

            result_text += f"{example['description']} ({example['distance']} км):\n"
            result_text += f"  Базовая стоимость: {cost:.2f} руб\n"
            if cost < min_price:
                result_text += f"  Применена минимальная цена: {final_cost:.2f} руб\n"
            else:
                result_text += f"  Итоговая стоимость: {final_cost:.2f} руб\n"
            result_text += "\n"

        QMessageBox.information(self, "Пример расчета", result_text)

    def show_export_dialog(self):
        """Показать диалог выбора фильтра для экспорта тарифов"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Экспорт тарифов в Excel")
        dialog.setMinimumWidth(400)

        layout = QVBoxLayout(dialog)

        # Группа с вариантами экспорта
        group_box = QGroupBox("Выберите что экспортировать:")
        group_layout = QVBoxLayout()

        self.export_options_group = QButtonGroup(dialog)

        options = [
            ("Все тарифы", "all"),
            ("Активные тарифы", "active"),
            ("Архивные тарифы", "archived"),
            ("Тарифы по типу груза", "by_cargo"),
            ("Бессрочные тарифы", "unlimited")
        ]

        for text, option_id in options:
            radio = QRadioButton(text)
            radio.option_id = option_id
            self.export_options_group.addButton(radio)
            group_layout.addWidget(radio)

        # Выбираем первый вариант по умолчанию
        if options:
            self.export_options_group.buttons()[0].setChecked(True)

        group_box.setLayout(group_layout)
        layout.addWidget(group_box)

        # Выбор типа груза (если выбрано "по типу груза")
        self.cargo_combo = QComboBox()
        self.cargo_combo.setVisible(False)
        if self.cargo_types:
            self.cargo_combo.addItems(self.cargo_types)

        def on_option_changed():
            selected_btn = self.export_options_group.checkedButton()
            if selected_btn and selected_btn.option_id == "by_cargo":
                self.cargo_combo.setVisible(True)
            else:
                self.cargo_combo.setVisible(False)

        self.export_options_group.buttonClicked.connect(on_option_changed)
        layout.addWidget(self.cargo_combo)

        # Кнопки
        buttons_layout = QHBoxLayout()
        export_btn = QPushButton("Экспорт")
        cancel_btn = QPushButton("Отмена")

        export_btn.clicked.connect(lambda: self.do_export(dialog))
        cancel_btn.clicked.connect(dialog.reject)

        buttons_layout.addWidget(export_btn)
        buttons_layout.addWidget(cancel_btn)
        layout.addLayout(buttons_layout)

        dialog.exec()

    def do_export(self, dialog):
        """Выполнить экспорт с выбранным фильтром"""
        # Получаем выбранный вариант
        selected_btn = self.export_options_group.checkedButton()
        if not selected_btn:
            QMessageBox.warning(dialog, "Ошибка", "Выберите вариант экспорта")
            return

        option_id = selected_btn.option_id

        # Дополнительные параметры
        extra_params = {}
        if option_id == "by_cargo":
            cargo_type = self.cargo_combo.currentText()
            if not cargo_type:
                QMessageBox.warning(dialog, "Ошибка", "Выберите тип груза")
                return
            extra_params["cargo_type"] = cargo_type

        # Получаем сессию
        session = self.get_session()
        if not session:
            QMessageBox.warning(dialog, "Ошибка",
                                "Не удалось получить доступ к базе данных")
            return

        # Создаем диалог для выбора файла
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        file_name, _ = QFileDialog.getSaveFileName(
            self,
            "Сохранить файл Excel",
            f"тарифы_{option_id}_{timestamp}.xlsx",
            "Excel files (*.xlsx)"
        )

        if not file_name:
            return  # Пользователь отменил

        try:
            # Экспортируем данные с фильтром
            if self.export_tariffs_to_excel(session, file_name, option_id, **extra_params):
                QMessageBox.information(
                    self,
                    "Экспорт завершен",
                    f"Данные успешно экспортированы в файл:\n{os.path.basename(file_name)}"
                )
                dialog.accept()
            else:
                QMessageBox.warning(dialog, "Ошибка",
                                    "Не удалось экспортировать данные")

        except Exception as e:
            QMessageBox.critical(dialog, "Ошибка экспорта",
                                 f"Произошла ошибка:\n{str(e)}")

    def get_session(self):
        """Получить сессию базы данных"""
        parent = self.parent()
        if parent and hasattr(parent, 'session'):
            return parent.session

        # Пробуем получить сессию из главного окна
        main_window = self.window()
        if main_window and hasattr(main_window, 'session'):
            return main_window.session

        return None

    def validate_and_accept(self):
        """Проверка данных и сохранение"""
        errors = []

        # Проверяем обязательные поля
        if not self.cargo_type_combo.currentText().strip():
            errors.append("Введите тип груза")

        if self.price_per_km_input.value() <= 0:
            errors.append("Цена за километр должна быть больше 0")

        if self.min_price_input.value() < 0:
            errors.append("Минимальная цена не может быть отрицательной")

        # Проверяем даты
        start_date = self.date_start_input.dateTime()
        end_date = self.date_end_input.dateTime() if not self.no_end_date_check.isChecked() else None

        if end_date and start_date > end_date:
            errors.append("Дата начала не может быть позже даты окончания")

        if errors:
            QMessageBox.warning(self, "Ошибка", "\n".join(errors))
            return

        self.accept()

    def get_data(self):
        """Получить данные из формы"""
        # Даты
        start_qdt = self.date_start_input.dateTime()
        start_date = datetime.datetime(
            start_qdt.date().year(),
            start_qdt.date().month(),
            start_qdt.date().day(),
            start_qdt.time().hour(),
            start_qdt.time().minute()
        )

        end_date = None
        if not self.no_end_date_check.isChecked():
            end_qdt = self.date_end_input.dateTime()
            end_date = datetime.datetime(
                end_qdt.date().year(),
                end_qdt.date().month(),
                end_qdt.date().day(),
                end_qdt.time().hour(),
                end_qdt.time().minute()
            )

        return {
            "cargo_type": self.cargo_type_combo.currentText().strip(),
            "price_per_km": self.price_per_km_input.value(),
            "min_price": self.min_price_input.value(),
            "date_start": start_date.isoformat(),
            "date_end": end_date.isoformat() if end_date else None,
            "description": self.description_input.toPlainText().strip()
        }

    def export_tariffs_to_excel(self, session, file_name, filter_type="all", **kwargs):
        """Экспорт тарифов в Excel с фильтрацией"""
        try:
            # Импортируем сервис тарифов
            from Services.Rate.services import (
                get_all_tariffs, get_active_tariffs,
                get_tariffs_by_cargo_type
            )
            import datetime

            # Получаем данные в зависимости от фильтра
            if filter_type == "all":
                tariffs_data = get_all_tariffs(session)
            elif filter_type == "active":
                tariffs_data = get_active_tariffs(session)
            elif filter_type == "archived":
                all_tariffs = get_all_tariffs(session)
                now = datetime.datetime.now()
                tariffs_data = [t for t in all_tariffs
                                if t.get("date_end") and
                                datetime.datetime.fromisoformat(t["date_end"]) < now]
            elif filter_type == "by_cargo":
                cargo_type = kwargs.get("cargo_type", "")
                tariffs_data = get_tariffs_by_cargo_type(session, cargo_type)
            elif filter_type == "unlimited":
                all_tariffs = get_all_tariffs(session)
                tariffs_data = [t for t in all_tariffs if t.get("date_end") is None]
            else:
                tariffs_data = get_all_tariffs(session)

            if not tariffs_data:
                QMessageBox.warning(self, "Нет данных",
                                    f"Нет тарифов по выбранному фильтру: {filter_type}")
                return False

            # Создаем Excel файл
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Тарифы"

            # Стили
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            title_font = Font(bold=True, size=14)
            title_alignment = Alignment(horizontal="center", vertical="center")

            # Заголовок
            ws.merge_cells('A1:H1')
            title_cell = ws['A1']
            title_cell.value = "ТАРИФЫ НА ПЕРЕВОЗКИ"
            title_cell.font = title_font
            title_cell.alignment = title_alignment

            # Заголовки столбцов
            headers = [
                ("ID", 10),
                ("Тип груза", 20),
                ("Цена за км (руб)", 15),
                ("Мин. цена (руб)", 15),
                ("Дата начала", 15),
                ("Дата окончания", 15),
                ("Статус", 12),
                ("Описание", 30)
            ]

            for col_idx, (header, width) in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            # Заполняем данными
            current_row = 4
            total_price_per_km = 0
            total_min_price = 0

            for tariff in tariffs_data:
                # Конвертируем даты в читаемый формат
                date_start = tariff.get("date_start", "")
                date_end = tariff.get("date_end", "")

                if date_start:
                    try:
                        dt = datetime.datetime.fromisoformat(date_start)
                        date_start_str = dt.strftime("%d.%m.%Y %H:%M")
                    except:
                        date_start_str = str(date_start)
                else:
                    date_start_str = ""

                if date_end:
                    try:
                        dt = datetime.datetime.fromisoformat(date_end)
                        date_end_str = dt.strftime("%d.%m.%Y %H:%M")
                    except:
                        date_end_str = str(date_end)
                else:
                    date_end_str = "Бессрочно"

                # Определяем статус
                is_active = tariff.get("is_active", False)
                status = "✅ Активен" if is_active else "⏸️ Архив"

                ws.cell(row=current_row, column=1, value=tariff.get("id", ""))
                ws.cell(row=current_row, column=2, value=tariff.get("cargo_type", ""))
                ws.cell(row=current_row, column=3, value=tariff.get("price_per_km", 0))
                ws.cell(row=current_row, column=4, value=tariff.get("min_price", 0))
                ws.cell(row=current_row, column=5, value=date_start_str)
                ws.cell(row=current_row, column=6, value=date_end_str)
                ws.cell(row=current_row, column=7, value=status)
                ws.cell(row=current_row, column=8, value=tariff.get("description", ""))

                # Форматирование для числовых полей
                price_cell = ws.cell(row=current_row, column=3)
                price_cell.number_format = '#,##0.00" руб"'

                min_price_cell = ws.cell(row=current_row, column=4)
                min_price_cell.number_format = '#,##0" руб"'

                # Собираем статистику
                total_price_per_km += tariff.get("price_per_km", 0)
                total_min_price += tariff.get("min_price", 0)

                current_row += 1

            # Добавляем границы для данных
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            for row in ws.iter_rows(min_row=3, max_row=current_row - 1, max_col=8):
                for cell in row:
                    cell.border = thin_border

            # Добавляем статистику
            stats_row = current_row + 2

            ws.cell(row=stats_row, column=1, value="СТАТИСТИКА").font = Font(bold=True)

            # Информация о фильтре
            filter_names = {
                "all": "Все тарифы",
                "active": "Активные тарифы",
                "archived": "Архивные тарифы",
                "by_cargo": f"Тарифы по типу груза: {kwargs.get('cargo_type', '')}",
                "unlimited": "Бессрочные тарифы"
            }

            ws.cell(row=stats_row + 1, column=1, value=f"Фильтр: {filter_names.get(filter_type, 'Все тарифы')}")
            ws.cell(row=stats_row + 1, column=2, value=f"Количество тарифов: {len(tariffs_data)}")

            if tariffs_data:
                avg_price_per_km = total_price_per_km / len(tariffs_data)
                avg_min_price = total_min_price / len(tariffs_data)

                ws.cell(row=stats_row + 2, column=1, value=f"Средняя цена за км: {avg_price_per_km:.2f} руб")
                ws.cell(row=stats_row + 2, column=2, value=f"Средняя минимальная цена: {avg_min_price:.2f} руб")

                # Подсчет активных и архивных
                active_count = sum(1 for t in tariffs_data if t.get("is_active", False))
                ws.cell(row=stats_row + 3, column=1, value=f"Активных тарифов: {active_count}")
                ws.cell(row=stats_row + 3, column=2, value=f"Архивных тарифов: {len(tariffs_data) - active_count}")

            # Добавляем дату экспорта
            ws.cell(row=stats_row + 5, column=1,
                    value=f"Экспортировано: {datetime.datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")

            # Автоподбор ширины для последних столбцов
            for col in range(5, 9):
                ws.column_dimensions[get_column_letter(col)].auto_size = True

            # Сохраняем файл
            wb.save(file_name)
            return True

        except ImportError as e:
            QMessageBox.critical(self, "Ошибка",
                                 f"Не удалось импортировать модули: {str(e)}")
            return False
        except Exception as e:
            print(f"Ошибка при экспорте тарифов: {str(e)}")
            return False