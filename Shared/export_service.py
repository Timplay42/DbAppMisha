# Services/Car/export_service.py
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from typing import List, Dict
from sqlalchemy.orm import Session


def export_cars_to_excel(session: Session, file_path: str, filters: Dict = None) -> bool:
    """
    Экспорт машин в Excel файл

    Args:
        session: SQLAlchemy сессия
        file_path: Путь для сохранения файла
        filters: Словарь с фильтрами (например, {"min_load_capacity": 10})

    Returns:
        bool: Успешность операции
    """
    try:
        from Services.Car.services import get_all_cars

        # Получаем все машины
        cars = get_all_cars(session)

        # Применяем фильтры если есть
        if filters:
            cars = apply_car_filters(cars, filters)

        # Создаем книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Машины"

        # Стили для заголовков
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Заголовки
        headers = ["ID", "Марка", "Государственный номер",
                   "Грузоподъемность (т)", "Тип кузова",
                   "Расход топлива (л/100км)", "Статус"]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Данные
        for row_idx, car in enumerate(cars, 2):
            # Определяем статус машины
            status = "Свободна"
            if hasattr(car, 'driver') and car.driver:
                status = f"Занята ({car.driver.full_name})"
            elif isinstance(car, dict):
                status = car.get("status", "Свободна")

            ws.append([
                car.get("id", row_idx - 1),
                car.get("brand", ""),
                car.get("license_plate", ""),
                car.get("load_capacity", 0),
                car.get("body_type", ""),
                car.get("fuel_consumption", 0),
                status
            ])

        # Автоподбор ширины столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Добавляем информацию о фильтрах
        if filters:
            ws.cell(row=ws.max_row + 2, column=1, value="Примененные фильтры:")
            for i, (key, value) in enumerate(filters.items(), 1):
                ws.cell(row=ws.max_row + 1, column=1, value=f"{key}: {value}")

        # Добавляем дату экспорта
        ws.cell(row=ws.max_row + 2, column=1,
                value=f"Экспортировано: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

        # Сохраняем файл
        wb.save(file_path)
        return True

    except Exception as e:
        print(f"Ошибка при экспорте машин: {str(e)}")
        return False


def apply_car_filters(cars: List[Dict], filters: Dict) -> List[Dict]:
    """Применение фильтров к списку машин"""
    filtered_cars = cars.copy()

    # Фильтр по грузоподъемности
    if "min_load_capacity" in filters:
        min_load = filters["min_load_capacity"]
        filtered_cars = [c for c in filtered_cars if c.get("load_capacity", 0) >= min_load]

    if "max_load_capacity" in filters:
        max_load = filters["max_load_capacity"]
        filtered_cars = [c for c in filtered_cars if c.get("load_capacity", 0) <= max_load]

    # Фильтр по типу кузова
    if "body_type" in filters:
        body_type = filters["body_type"].lower()
        filtered_cars = [c for c in filtered_cars if body_type in c.get("body_type", "").lower()]

    # Фильтр по марке
    if "brand" in filters:
        brand = filters["brand"].lower()
        filtered_cars = [c for c in filtered_cars if brand in c.get("brand", "").lower()]

    # Фильтр по расходу топлива
    if "max_fuel_consumption" in filters:
        max_fuel = filters["max_fuel_consumption"]
        filtered_cars = [c for c in filtered_cars if c.get("fuel_consumption", 0) <= max_fuel]

    # Фильтр по статусу (с водителем/без)
    if "has_driver" in filters:
        has_driver = filters["has_driver"]
        if has_driver:
            filtered_cars = [c for c in filtered_cars if c.get("has_driver", False)]
        else:
            filtered_cars = [c for c in filtered_cars if not c.get("has_driver", True)]

    return filtered_cars


def get_car_export_options() -> List[Dict]:
    """Получить варианты экспорта для машин"""
    return [
        {"id": "all", "name": "Все машины", "description": "Экспорт всех машин"},
        {"id": "free", "name": "Свободные машины", "description": "Машины без назначенного водителя"},
        {"id": "heavy", "name": "Грузоподъемные машины", "description": "Грузоподъемность > 10 тонн"},
        {"id": "efficient", "name": "Экономичные машины", "description": "Расход топлива < 15 л/100км"}
    ]